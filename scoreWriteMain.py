from browsermobproxy import Server
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import json
import requests
import random
import openpyxl
import os
import ddddocr
import cv2
import numpy as np
import uuid
from openpyxl.utils import get_column_letter
from getExcelValue import read_excel, write_excel
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from openpyxl.styles import PatternFill
import sys

# 从Excel表格中读取账号和密码 第2行开始
current_row = 2
accountInfo = None
reason = ""
fileUrl = "D:\getScore\蔡老师-学员成绩表1月5日.xlsx"


def get_resource_path(relative_path):
    if getattr(sys, "frozen", False):
        # The application is frozen
        base_path = sys._MEIPASS
    else:
        # The application is not frozen
        base_path = os.path.dirname(os.path.abspath(__file__))
    print("驱动包path:" + os.path.join(base_path, relative_path))
    return os.path.join(base_path, relative_path)


def getCJMC(cj):
    cjArr = [
        {"cjdm": "++", "mc": "满分"},
        {"cjdm": "+-", "mc": "免考"},
        {"cjdm": "-0", "mc": "取消成绩"},
        {"cjdm": "+1", "mc": "及格"},
        {"cjdm": "+2", "mc": "中等"},
        {"cjdm": "+3", "mc": "良好"},
        {"cjdm": "+4", "mc": "优秀"},
        {"cjdm": "+0", "mc": "不及格"},
    ]

    for item in cjArr:
        if item["cjdm"] == cj:
            return item["mc"]

    return cj


# 写入密码错误等信息
def writeErrorInfo(ksh, errorInfo):
    workbook = openpyxl.load_workbook(fileUrl)
    match_sheet = None
    match_row = None
    # 遍历所有sheet页
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # 查找包含 "准考证号" 的列
        ksh_column = None
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col).value
            if header_cell and "准考证号" in str(header_cell):
                ksh_column = col
                break

        if ksh_column is None:
            print(f'"准考证号" 列不存在: {sheet_name}')
            continue

        # 查找匹配的行
        for row in range(2, sheet.max_row + 1):  # 从第二行开始遍历，假设第一行为标题行
            cell_value = sheet.cell(row=row, column=ksh_column).value
            if str(cell_value).strip() == str(ksh).strip():
                match_sheet = sheet
                match_row = row
                break

        if match_sheet and match_row:
            break
    if not match_sheet or not match_row:
        print(f"未找到准考证号为 {ksh} 的记录")
    else:
        # print(f"找到匹配学生在 sheet: {match_sheet.title}, 行: {match_row}")
        # 根据查到的写入查询明细
        # 写入成绩match_sheet和match_row 再第一个单元格写入详情
        match_sheet.cell(row=match_row, column=1).value = errorInfo
        workbook.save((fileUrl))


def writeExcel(resultData):
    begin_time = time.time()
    # result_str = json.dumps(resultData, ensure_ascii=False, indent=4)
    # print("最终结果:\n" + result_str)
    print("写入excel中...")

    # 加载Excel工作簿
    # 打印当前工作目录
    # print(f"Current working directory: {os.getcwd()}")

    # 获取当前脚本所在的目录
    # script_dir = os.path.dirname(os.path.abspath(__file__))
    # 构建文件的完整路径
    # file_path = os.path.join(script_dir, '彭老师学员成绩表1-9.xlsx')
    workbook = openpyxl.load_workbook(fileUrl)
    # 获取准考证号
    ksh = resultData["ksh"]
    xm = resultData["xm"]
    phone = resultData["phone"]
    idcard = resultData["idcard"]
    kmcjTotal = resultData["kmcjTotal"]
    sjkcjTotal = resultData["sjkcjTotal"]
    detailStr = (
        "笔试已过" + str(kmcjTotal) + "科" + ",实践课已过" + str(sjkcjTotal) + "科"
    )
    match_sheet = None
    match_row = None

    # 遍历所有sheet页
    for sheet_name in workbook.sheetnames:
        find_time = time.time()
        sheet = workbook[sheet_name]

        # 查找包含 "准考证号" 的列
        ksh_column = None
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col).value
            if header_cell and "准考证号" in str(header_cell):
                ksh_column = col
                # print(
                #     f'找到 "准考证号" 在{sheet_name},第: {openpyxl.utils.get_column_letter(col)}列耗时:{time.time() - find_time:.4f} 秒'
                # )
                break

        if ksh_column is None:
            print(f'"准考证号" 列不存在: {sheet_name}')
            continue

        # print(f"sheet.max_row: ", sheet.max_row)
        # 查找匹配的行
        for row in range(2, sheet.max_row + 1):  # 从第二行开始遍历，假设第一行为标题行
            cell_value = sheet.cell(row=row, column=ksh_column).value
            if str(cell_value).strip() == str(ksh).strip():
                # print('找到准考证号为 " + str(ksh) + " 的记录"')
                match_sheet = sheet
                match_row = row
                # print(f"找到匹配的行耗时:{time.time() - find_time:.4f} 秒")
                break

        if match_sheet and match_row:
            break
    if not match_sheet or not match_row:
        # print(f"未找到准考证号为 {ksh} 的记录")
        reason = "未找到准考证号为" + str(ksh) + "的记录"
        write_excel(accountInfo.row, reason)
        driver.back()
        driver.refresh()
        return

    # print(f"匹配准考证号耗时:{time.time() - begin_time:.4f} 秒")
    # print(f"找到匹配学生在 sheet: {match_sheet.title}, 行: {match_row}")

    # 根据查到的写入查询明细
    # 写入成绩match_sheet和match_row 再第一个单元格写入详情
    match_sheet.cell(row=match_row, column=1).value = detailStr
    # match_sheet.cell(row=match_row, column=2).value = xm 注释先，不要姓名匹配了

    # 获取第一行的所有单元格内容
    first_row_cells = []
    for col in range(1, match_sheet.max_column + 1):
        cell_value = match_sheet.cell(row=1, column=col).value
        if cell_value is not None:
            first_row_cells.append(cell_value.strip())
        else:
            first_row_cells.append(None)

    # for index, title in enumerate(first_row_cells): 注释先，不要姓名匹配了
    #     if title is not None and title.strip() == "姓名":
    #         old_name = match_sheet.cell(row=match_row, column=index + 1).value
    #         match_sheet.cell(row=match_row, column=3).value = old_name == xm
    #         if old_name != xm:
    #             color_fill = PatternFill(
    #                 start_color="FF0000", end_color="FF0000", fill_type="solid"
    #             )
    #         else:
    #             color_fill = PatternFill(
    #                 start_color="00FF00", end_color="00FF00", fill_type="solid"
    #             )
    #         match_sheet.cell(row=match_row, column=3).fill = color_fill

    # 创建 KMDM 到列索引的映射（模糊匹配）
    kmdm_to_col_index = {}
    for index, title in enumerate(first_row_cells):
        if title is not None:
            for item in resultData["rows"]:
                if item["KMDM"] is not None and (
                    item["KMDM"] in title or title in item["KMDM"]
                ):
                    kmdm_to_col_index[item["KMDM"]] = index
                    break

    for item in resultData["rows"]:
        kmdm = item["KMDM"]
        cj = item["CJ"]

        if kmdm in kmdm_to_col_index:
            col_index = kmdm_to_col_index[kmdm]
            cell = match_sheet.cell(
                row=match_row, column=col_index + 1
            )  # 注意这里需要加1，因为openpyxl的列索引是从1开始的
            cell.value = getCJMC(cj)
            # print(f"科目匹配成功 {kmdm}{item['KMMC']}:{getCJMC(cj)} ")
        else:
            print(f"未找到科目代码 {kmdm} 在 Excel 中")

    # 保存工作簿
    workbook.save(fileUrl)
    print(xm + "成绩已成功写入 Excel 文件")
    reason = (
        "找到excel匹配学生在 sheet: "
        + str(match_sheet.title)
        + ", 第"
        + str(match_row)
        + " 行,成绩已成功写入 Excel 文件"
    )
    write_excel(accountInfo.row, reason)
    print(f"写入excel耗时:{time.time() - begin_time:.4f} 秒")
    driver.back()
    driver.refresh()
    goLogin()


def getRequest(Cookies):
    # k考生信息
    ksxxUrl = "https://www.eeagd.edu.cn/zkselfec/gkcx/queryKsxx.jsmeb"
    # 已通过笔试成绩
    kmcjUrl = "https://www.eeagd.edu.cn/zkselfec/gdbk/queryKmcj.jsmeb"
    # 已通过实践课成绩
    sjkcjUrl = "https://www.eeagd.edu.cn/zkselfec/gdbk/querySjkcj.jsmeb"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Content-Type": "application/json;charset=UTF-8",
        "Host": "www.eeagd.edu.cn",
        "Cookie": Cookies,
    }
    payload = {}
    response = requests.post(ksxxUrl, headers=headers, json=payload)
    # ("Response Body:", response.text)
    ksxx_json = json.loads(response.text)
    if "error" in ksxx_json:
        error_code = ksxx_json["error"]["code"]
        error_message = ksxx_json["error"]["message"]
        print(f"姓名: {accountInfo.name}")
        print(f"错误代码: {error_code}, 错误消息: {error_message}")

        return  # 如果是在函数内部，使用 return 跳出函数
    else:
        print(
            f"当前考生姓名: {ksxx_json['result']['data']['xm']},准考证号：{ksxx_json['result']['data']['ksh']}"
        )
    # print("Response Headers:", response.headers)
    # print("Response Body:", response.text)

    payload = [{"page": 1, "rows": 100}, "", ""]
    response = requests.post(kmcjUrl, headers=headers, json=payload)
    if response.text:
        kmcj_json = json.loads(response.text)
        # print(f"已通过笔试科目: {kmcj_json['result']['total']},科目信息{kmcj_json['result']['rows']}")
        print(f"已通过笔试科目: {kmcj_json['result']['total']}")
    else:
        print(f"未查询到已通过笔试科目")
    # print("Status Code:", response.status_code)
    # print("Response Headers:", response.headers)

    response = requests.post(sjkcjUrl, headers=headers, json=payload)
    # print("Response Body:", response.text)
    if response.text:
        sjkcj_json = json.loads(response.text)
        print(f"已通过实践科目: {sjkcj_json['result']['total']}")
    else:
        print(f"未查询到已通过实践科目")
    # 合并笔试成绩和实践成绩结果
    # 合并 result.rows 数组
    merged_rows = kmcj_json["result"]["rows"] + sjkcj_json["result"]["rows"]

    # 累加 result.total
    merged_total = int(kmcj_json["result"]["total"]) + int(
        sjkcj_json["result"]["total"]
    )

    # 构建最终的 JSON 对象
    merged_result = {
        "xm": ksxx_json["result"]["data"]["xm"],
        "ksh": ksxx_json["result"]["data"]["ksh"],
        "total": merged_total,
        "kmcjTotal": kmcj_json["result"]["total"],
        "sjkcjTotal": sjkcj_json["result"]["total"],
        "rows": merged_rows,
        "idcard": ksxx_json["result"]["data"]["zjdm"],
        "phone": ksxx_json["result"]["data"]["yddh"],
    }
    # 调用写入excel函数
    writeExcel(merged_result)


# 启动 BrowserMob Proxy 服务器
server = Server(get_resource_path(r"browsermob-proxy-2.1.4\bin\browsermob-proxy.bat"))
server.start()
proxy = server.create_proxy()
print(proxy.proxy)
print(proxy.port)
first_start_time = time.time()

# 配置 Selenium WebDriver 使用代理
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument(f"--proxy-server={proxy.proxy}")

# 指定 ChromeDriver 路径
driver_path = get_resource_path(r"chromedriver-win64\chromedriver.exe")

# 初始化WebDriver
driver = webdriver.Chrome(service=Service(driver_path), options=chrome_options)

# 目标 URL
login_url = "https://www.eeagd.edu.cn/zkselfec/login/login.jsp"


# 自动获取excel表格的账号密码登录
def goLogin():
    global reason
    global accountInfo
    global current_row
    accountInfo = read_excel(current_row)
    if accountInfo is None:
        print("成绩录入完成，共" + str(current_row - 2) + "条数据")
        print(f"整体耗时:{time.time() - first_start_time:.4f} 秒")
        return

    print(f"当前查询到第{current_row-1}条数据")
    # 只要读取到账号密码 就+1 下次就找下一行
    reason = ""  # 清空查询明细
    current_row += 1
    start_time = time.time()
    # 查找验证码图片元素
    captcha_element = driver.find_element(
        By.XPATH, '//*[@id="loginForm"]/div[3]/div/a/img'
    )  # 根据实际情况修改定位方式

    # 获取验证码图片的位置和大小
    location = captcha_element.location
    size = captcha_element.size
    # print(f"查找验证码位置耗时:{time.time() - start_time:.4f} 秒")

    start_time = time.time()
    # 截取整个屏幕
    screenshot = driver.get_screenshot_as_png()
    screenshot_np = np.frombuffer(screenshot, np.uint8)
    screenshot_cv = cv2.imdecode(screenshot_np, cv2.IMREAD_COLOR)
    # print(f"截取整个屏幕耗时: {time.time() - start_time:.4f} 秒")

    # 计算验证码图片的实际位置
    start_time = time.time()
    left = int(location["x"])
    top = int(location["y"])
    right = int(left + size["width"])
    bottom = int(top + size["height"])
    # print(f"计算验证码图片的实际位置耗时: {time.time() - start_time:.4f} 秒")

    # 截取验证码图片
    start_time = time.time()
    captcha_image = screenshot_cv[top:bottom, left:right]
    # print(f"截取验证码图片耗时: {time.time() - start_time:.4f} 秒")

    # 保存原始验证码图片（可选）cl
    # 开始计时
    # 生成一个 UUID
    unique_id = uuid.uuid4()
    filename = f"{unique_id}.png"
    captcha_save_path = "D:\\getScore\\captchaImg\\{}.png".format(
        unique_id
    )  # 使用双反斜杠进行转义  # 路径不能包含中文
    # 检查文件夹是否存在
    # folder_path = os.path.dirname(captcha_save_path)
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)
    #     print(f"文件夹 {folder_path} 不存在，已创建。")
    # else:
    #     print(f"文件夹 {folder_path} 存在。")

    # 打印保存路径
    # print(f"验证码保存路径: {captcha_save_path}")
    start_time = time.time()
    cv2.imwrite(captcha_save_path, captcha_image)
    # 结束计时
    # print(f"验证码图片保存耗时: {time.time() - start_time:.4f} 秒")

    # 检查文件是否存在
    if not os.path.exists(captcha_save_path):
        raise FileNotFoundError(f"文件未成功保存: {captcha_save_path}")

    start_time = time.time()
    ocr_model_path = get_resource_path(r"onnx\common_old.onnx")
    ocr = ddddocr.DdddOcr(show_ad=False, charsets_path=ocr_model_path)
    # ocr = ddddocr.DdddOcr()  # 切换为第二套ocr模型为 ocr = ddddocr.DdddOcr(beta=True)
    # print(f"初始化ddddocr耗时: {time.time() - start_time:.4f} 秒")

    start_time = time.time()
    with open(captcha_save_path, "rb") as file:
        image = file.read()
    captcha_result = ocr.classification(image)
    print("验证码识别结果：" + captcha_result)
    print(f"识别验证码耗时: {time.time() - start_time:.4f} 秒")

    # 输入账号
    username_input = driver.find_element(By.ID, "username")
    username_input.send_keys(accountInfo.account)  # 替换为实际的用户名
    # time.sleep(0.5)

    # 输入密码
    password_input = driver.find_element(By.ID, "password")
    password_input.send_keys(accountInfo.password)  # 替换为实际的密码
    # time.sleep(0.5)

    # 输入验证码
    captcha_input = driver.find_element(By.ID, "addcode")
    captcha_input.clear()  # 清空验证码输入框（如果有默认值）
    captcha_input.send_keys(captcha_result)
    # time.sleep(0.5)

    # 点击登录按钮
    login_button = driver.find_element(By.ID, "logincommit")
    login_button.click()


# 打开登录页面 并且自动登录
start_time = time.time()
driver.get(login_url)

# 增加等待时间以确保页面完全加载
time.sleep(3)  # 根据实际情况调整等待时间
# print(f"打开登录界面耗时:{time.time() - start_time:.4f} 秒")
# 开始捕获流量
proxy.new_har("fj", options={"captureHeaders": True, "captureContent": True})
goLogin()
# 实时监控并打印请求和响应
try:
    while True:
        try:
            alert = driver.switch_to.alert
            print("检测到界面提示信息:", alert.text)
            if "验证码错误" in alert.text:
                current_row -= 1
            if "用户名或密码错误" in alert.text:
                # 写入成绩表的查询结果明细
                write_excel(accountInfo.row, alert.text)
                writeErrorInfo(accountInfo.account, alert.text)
            alert.accept()  # 自动点击确认
            time.sleep(0.5)
            goLogin()
        except Exception as e:
            # 没有找到alert，继续循环
            pass
        # 获取当前捕获的 HAR 数据
        har = proxy.har
        if har and "log" in har and "entries" in har["log"]:
            for entry in har["log"]["entries"]:
                request_url = entry["request"]["url"]
                # 校验登录是否成功
                if (
                    request_url
                    == "https://www.eeagd.edu.cn/zkselfec/gkcx/queryKsxx.jsmeb"
                    # or request_url
                    # == "https://www.eeagd.edu.cn/zkselfec/gkcx/queryKsxx.jsmeb"
                    # or request_url
                    # == "https://www.eeagd.edu.cn/zkselfec/user/getRoleMenus.jsmeb"
                ):
                    # 自考网站进入登录界面自动分配cookie 登录成功则cookie生效，否则不生效
                    # 所以这直接拿cookie请求考生信息接口即可，请求成功就登录成功  否则就失败
                    print(f"登录成功获取考生信息...: {request_url}")
                    # 特别打印 Cookie
                    cookie_header = next(
                        (
                            header
                            for header in entry["request"]["headers"]
                            if header["name"].lower() == "cookie"
                        ),
                        None,
                    )
                    if cookie_header and cookie_header["value"]:
                        print(f"获取到Cookie: {cookie_header['value']}")
                        # 等待1秒 自考后端登录需要时间
                        time.sleep(0.5)
                        # 构建Post请求查询考生信息和已通过课程
                        getRequest(cookie_header["value"])
                    else:
                        print("没找到有效Cookie")
                        continue
            # 使用 new_har 方法重新开始一个新的 HAR 会话
            proxy.new_har(
                "fj", options={"captureHeaders": True, "captureContent": True}
            )

            # 添加一个小延迟以避免 CPU 占用过高
            time.sleep(1)

except KeyboardInterrupt:
    # 用户按下 Ctrl+C 停止程序
    print("程序已停止")

finally:
    # 关闭 WebDriver 和 BrowserMob Proxy 服务器
    print("程序关闭中...")
    driver.quit()
    server.stop()
    print("程序关闭成功")
