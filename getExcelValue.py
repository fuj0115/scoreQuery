import openpyxl


class AccountInfo:
    def __init__(self, row, name, account, password):
        self.row = row
        self.name = name
        self.account = account
        self.password = password

    def __repr__(self):
        return f"账号信息(row={self.row},name={self.name},account={self.account}, password={self.password})"


def read_excel(row):
    # 定义文件路径
    file_path = r"D:\getScore\账号密码.xlsx"

    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择活动的工作表
    sheet = workbook.active

    try:
        # 获取指定行的第二列和第三列的值
        name = sheet.cell(row=row, column=1).value
        account = sheet.cell(row=row, column=2).value
        password = sheet.cell(row=row, column=3).value
        if account is None or password is None:
            return None
        account_info = AccountInfo(row, name, account, password)
        print(f"账号: {account}, 密码: {password}")
        return account_info

    except Exception as e:
        print(f"发生错误: {e}")


def write_excel(row, reason):
    # 定义文件路径
    file_path = r"D:\getScore\账号密码.xlsx"

    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择活动的工作表
    sheet = workbook.active

    try:
        # 在第四列写入“读取成功”
        sheet.cell(row=row, column=4).value = reason

        # 保存工作簿
        workbook.save(file_path)

        print("查询信息已写入:账号密码.xlsx")

    except Exception as e:
        print(f"发生错误: {e}")


# 示例调用
if __name__ == "__main__":
    print(read_excel(2).account)
